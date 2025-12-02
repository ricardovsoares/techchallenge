import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import re


def criar_pasta_saida(caminho_pasta="dados_exportados"):
    """
    Cria a pasta de saída se ela não existir.

    Args:
        caminho_pasta: Caminho da pasta a ser criada (padrão: dados_exportados)

    Returns:
        str: Caminho completo da pasta criada
    """
    try:
        if not os.path.exists(caminho_pasta):
            os.makedirs(caminho_pasta)
            print(f"✓ Pasta criada: {caminho_pasta}")
        return caminho_pasta
    except Exception as e:
        print(f"Erro ao criar pasta: {e}")
        return None


def gerar_nome_arquivo_versionado(caminho_completo):
    """
    Gera um nome de arquivo versionado se o arquivo já existir.
    Adiciona sufixo numérico: arquivo.xlsx → arquivo_001.xlsx → arquivo_002.xlsx

    Args:
        caminho_completo: Caminho completo do arquivo desejado

    Returns:
        str: Caminho com nome versionado se necessário
    """
    if not os.path.exists(caminho_completo):
        return caminho_completo

    # Separa caminho, nome e extensão
    diretorio = os.path.dirname(caminho_completo)
    nome_arquivo = os.path.basename(caminho_completo)
    nome_base, extensao = os.path.splitext(nome_arquivo)

    # Verifica se já tem sufixo numérico
    match = re.match(r'^(.+?)_(\d+)$', nome_base)
    if match:
        nome_base = match.group(1)

    # Encontra o próximo número disponível
    contador = 1
    while True:
        novo_nome = f"{nome_base}_{contador:03d}{extensao}"
        novo_caminho = os.path.join(diretorio, novo_nome)

        if not os.path.exists(novo_caminho):
            return novo_caminho

        contador += 1


def salvar_em_excel(resultados, caminho_pasta="dados_exportados", nome_arquivo="produtos.xlsx", auto_versionar=True):
    """
    Salva os resultados em um arquivo Excel com formatação profissional.

    Args:
        resultados: Lista com dicionários de produtos
        caminho_pasta: Caminho da pasta onde será salvo (padrão: dados_exportados)
        nome_arquivo: Nome do arquivo a ser criado (padrão: produtos.xlsx)
        auto_versionar: Se True, cria versões numeradas de arquivos duplicados (padrão: True)

    Returns:
        bool: True se salvo com sucesso, False caso contrário
    """
    try:
        if not resultados:
            print("⚠ Nenhum dado para salvar.")
            return False

        # Cria a pasta de saída
        pasta = criar_pasta_saida(caminho_pasta)
        if pasta is None:
            return False

        # Caminho completo do arquivo
        caminho_completo = os.path.join(pasta, nome_arquivo)

        # Aplicar versionamento se ativado
        if auto_versionar:
            caminho_completo = gerar_nome_arquivo_versionado(caminho_completo)
            nome_arquivo_final = os.path.basename(caminho_completo)
        else:
            nome_arquivo_final = nome_arquivo

        # Cria um DataFrame a partir dos resultados
        df = pd.DataFrame(resultados)

        # Renomeia as colunas com nomes mais legíveis
        df = df.rename(columns={
            'url': 'url',
            'titulo': 'titulo',
            'descricao': 'descricao',
            'preco': 'preco',
            'rating': 'rating',
            'disponibilidade': 'disponibilidade',
            'categoria': 'categoria',
            'imagem_url': 'imagem'
        })

        # Adiciona coluna de índice
        df.insert(0, 'id', range(1, len(df) + 1))

        # Cria o writer do Excel
        with pd.ExcelWriter(caminho_completo, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Produtos', index=False)

            # Acessa o workbook para formatação
            workbook = writer.book
            worksheet = writer.sheets['Produtos']

            # Formata o header
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")

            # Ajusta larguras das colunas
            worksheet.column_dimensions['A'].width = 5     # ID
            worksheet.column_dimensions['B'].width = 50    # URL
            worksheet.column_dimensions['C'].width = 50    # Titulo
            worksheet.column_dimensions['D'].width = 150   # Descrição
            worksheet.column_dimensions['E'].width = 10    # Preço
            worksheet.column_dimensions['F'].width = 10    # Rating
            worksheet.column_dimensions['G'].width = 10    # Disponibilidade
            worksheet.column_dimensions['H'].width = 20    # Categoria
            worksheet.column_dimensions['I'].width = 50    # Imagem

            # Adiciona bordas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=9):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Congela a linha de header
            worksheet.freeze_panes = "A2"

        # Exibe mensagem de sucesso
        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        print(f"\n{'='*70}")
        print(f"✓ ARQUIVO SALVO COM SUCESSO")
        print(f"{'='*70}")
        print(f"Caminho: {caminho_completo}")
        print(f"Nome do arquivo: {nome_arquivo_final}")
        print(f"Total de produtos: {len(df)}")
        print(f"Data/Hora: {timestamp}")
        print(f"{'='*70}\n")

        return True

    except Exception as e:
        print(f"\nErro ao salvar Excel: {e}")
        return False


def salvar_em_csv(resultados, caminho_pasta="dados_exportados", nome_arquivo="produtos.csv", auto_versionar=True):
    """
    Salva os resultados em um arquivo CSV (alternativa ao Excel).

    Args:
        resultados: Lista com dicionários de produtos
        caminho_pasta: Caminho da pasta onde será salvo (padrão: dados_exportados)
        nome_arquivo: Nome do arquivo a ser criado
        auto_versionar: Se True, cria versões numeradas de arquivos duplicados (padrão: True)

    Returns:
        bool: True se salvo com sucesso, False caso contrário
    """
    try:
        if not resultados:
            print("⚠ Nenhum dado para salvar.")
            return False

        # Cria a pasta de saída
        pasta = criar_pasta_saida(caminho_pasta)
        if pasta is None:
            return False

        # Caminho completo do arquivo
        caminho_completo = os.path.join(pasta, nome_arquivo)

        # Aplicar versionamento se ativado
        if auto_versionar:
            caminho_completo = gerar_nome_arquivo_versionado(caminho_completo)
            nome_arquivo_final = os.path.basename(caminho_completo)
        else:
            nome_arquivo_final = nome_arquivo

        df = pd.DataFrame(resultados)
        df = df[['url', 'titulo', 'descricao', 'preco', 'rating',
                 'disponibilidade', 'categoria', 'imagem_url']]

        df = df.rename(columns={
            'url': 'url',
            'titulo': 'titulo',
            'descricao': 'descricao',
            'preco': 'preco',
            'rating': 'rating',
            'disponibilidade': 'disponibilidade',
            'categoria': 'categoria',
            'imagem_url': 'imagem'
        })

        df.insert(0, 'id', range(1, len(df) + 1))
        df.to_csv(caminho_completo, index=False, encoding='utf-8')

        print(f"\n✓ Arquivo CSV salvo com sucesso")
        print(f"  Caminho: {caminho_completo}")
        print(f"  Nome: {nome_arquivo_final}")
        print(f"  Total de produtos: {len(df)}\n")

        return True

    except Exception as e:
        print(f"\nErro ao salvar CSV: {e}")
        return False
